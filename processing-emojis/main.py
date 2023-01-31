from openpyxl import load_workbook
from processing_text import processing_text
import csv

filename = "data.xlsx"
smile_csv = "smile.csv"
min_row = 2

def main():
  try:
    # 166 smile emojis
    smile_doc = open(smile_csv, "r", newline="")
    reader = csv.reader(smile_doc, delimiter=',')
    smile_emoji_array = []
    for row in reader:
      smile_emoji_array.append(row)    
    # write xlsx
    workbook = load_workbook(filename=filename)
    sheet = workbook.active
    index = min_row
    for value in sheet.iter_rows(min_row=min_row, values_only=True):      
      if (value[0] != None):
        text = processing_text(value[0], smile_emoji_array)        
        emoji_all = ""
        if (len(text["emoji_all"]) > 0):
          for emoji in text["emoji_all"]:
            emoji_all += emoji        
        sheet[f"B{index}"] = emoji_all
        sheet[f"C{index}"] = str(text["count_emoji_all"])
        sheet[f"D{index}"] = str(text["emj_emotion"])
        sheet[f"E{index}"] = str(text["text_without_emoji"])
      index += 1
    workbook.save(filename=filename)
  except Exception as e:
    print(e.args)

if __name__ == "__main__":
  main()
